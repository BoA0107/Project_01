# -*- coding:utf-8 -*-
from openpyxl import *


# read excel


def read_doc(filename, sheetname, MaxLine=0):
    wb = load_workbook(filename)
    sheet = wb.get_sheet_by_name(sheetname)
    y = sheet.max_column
    x = sheet.max_row

    content = []
    if MaxLine == 0:
        for i in range(1, x + 1):
            c = []
            for j in range(1, y + 1):
                c.append(sheet.cell(i, j).value)
            content.append(c)
    else:
        c = []
        for j in range(1, y + 1):
            c.append(sheet.cell(1, j).value)
        content.append(c)
        for i in range(x + 1 - MaxLine, x + 1):
            c = []
            for j in range(1, y + 1):
                c.append(sheet.cell(i, j).value)
            content.append(c)
    return content


def make_dct(file):
    dct = {}
    for i in range(len(file)):
        x, y = file[i]
        if (x != None) and (y == None):
            key = x
            dct[key] = {}
            for j in range(i + 1, len(file)):
                x, y = file[j]
                if y != None:
                    dct[key][x] = y
                else:
                    break
    return dct


def read_line(filename, sheetname):
    wb = load_workbook(filename)
    sheet = wb.get_sheet_by_name(sheetname)
    content = []
    x = sheet.max_row
    for i in range(1, x + 1):
        content.append(sheet.cell(i, 1).value)
    return content


def read_dict(filename, sheetname):
    wb = load_workbook(filename)
    sheet = wb.get_sheet_by_name(sheetname)
    x = sheet.max_row
    dct = {}
    for i in range(1, x + 1):
        content = sheet.cell(i, 1).value
        if content != None:
            key = content
            dct[key] = []
            val = sheet.cell(i, 2).value
            dct[key].append(val)
        else:
            val = sheet.cell(i, 2).value
            dct[key].append(val)
    return dct
